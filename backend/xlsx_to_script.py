#!/usr/bin/env python3
"""Convert a "CAN Test Script Editor" xlsx (see CAN_Test_Script_Editor_Rev01.xlsx)
into the JSON step-script format used by CAN_simulator's test runner (see
test_runner_service.py).

This module lives in backend/ so the FastAPI upload endpoint can import it
directly (`from xlsx_to_script import convert, ScriptError`) to accept .xlsx
uploads in addition to .json. It can still be run as a standalone CLI for
development/validation:
    python backend/xlsx_to_script.py samples/CAN_Test_Script_Editor_Rev01.xlsx
    python backend/xlsx_to_script.py in.xlsx -o out.json --sheet Script

Column layout of the "Script" sheet, one physical row per step, data rows
start right after the row whose column B literally reads "Step 종류 선택":
    A: always "type" (fixed label, unused)
    B: step type -- ID / Power / delay / CANReq / CANEv / CANResp / Audio / loop / "]"
    C: a label word for most types (num/command/ms/cycle) -- but for
       CANReq/CANEv/CANResp it instead holds a *duplicate* CAN signal name,
       an artifact of the sheet's own live search box. It is never a JSON
       field on its own and is ignored here; the real Message/Signal values
       for those rows come from columns D and F.
    D: value 1 -- num / command / ms / loop cycle count / CAN Message name
    E: label 2 (unused)
    F: value 2 -- ID's cycle count / CAN Signal name
    G: label 3 (unused, "Value" on CAN rows)
    H: value 3 -- CAN Value (hex string, e.g. "0x01")
    I, J: DBC reference text for the spreadsheet's own dropdown UI -- ignored.

loop/"]" rows are a flat bracket-matched pair (like "(" ")"), not the nested
{"type": "loop", "cycle": N, "steps": [...]} block the JSON format uses, so
loop bodies are collected on a stack and only nested into a "loop" step once
the matching "]" row closes them. This also makes nested loops work, even
though this template's own dropdown doesn't offer nesting.

CANMsgS/CANMsgE rows are a similar flat bracket pair, marking a run of
CANReq/CANEv rows that all target the same CAN message and must be sent as
one frame with multiple signals -- e.g.:
    CANMsgS
    CANReq | ... | CLU_AMP_01_200ms | Signal | Warn_Sound_ETC    | Value | 0x01
    CANReq | ... | CLU_AMP_01_200ms | Signal | Warn_Sound_TikTok | Value | 0x02
    CANMsgE
becomes a single step:
    {"type": "CANReq", "Message": "CLU_AMP_01_200ms", "Signals": [
        {"Signal": "Warn_Sound_ETC", "Value": "0x01"},
        {"Signal": "Warn_Sound_TikTok", "Value": "0x02"}
    ]}
All rows in the block must share the same step type (CANReq or CANEv -- not
CANResp, which has no "Signals" equivalent in the backend) and the same
Message; a mismatch is an error. Groups don't nest inside loops or each
other.

Note: the ID row's repeat count must be written as "Cycle" (capital C) to
match test_runner_service.py's case-boundary parser -- the sheet's own
auto-label for that column is lowercase "cycle", so this is corrected here
rather than copied verbatim.

Usage:
    pip install openpyxl
    python backend/xlsx_to_script.py samples/CAN_Test_Script_Editor_Rev01.xlsx
    python backend/xlsx_to_script.py in.xlsx -o out.json --sheet Script
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Optional

import openpyxl

HEADER_MARKER = "Step 종류 선택"  # column B of the header row in the Script sheet
STEP_TYPES = {"ID", "Power", "delay", "CANReq", "CANEv", "CANResp", "Audio", "loop"}
CAN_TYPES = {"CANReq", "CANEv", "CANResp"}
GROUPABLE_CAN_TYPES = {"CANReq", "CANEv"}  # CANResp has no "Signals" equivalent
LOOP_END = "]"
MSG_GROUP_START = "CANMsgS"
MSG_GROUP_END = "CANMsgE"


class ScriptError(ValueError):
    """A row in the Script sheet couldn't be converted."""


def _require(value, row: int, field: str):
    if value is None or (isinstance(value, str) and value.strip() == ""):
        raise ScriptError(f"{row}행: 필수 값 {field}가 비어 있습니다")
    return value


def _num_str(value) -> str:
    """1 / 1.0 / "1" -> "1", matching test_script_Rev01.json's num convention."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _build_step(row: int, b: str, d, f, h) -> dict:
    if b == "ID":
        num = _require(d, row, "D열(num)")
        cycle = _require(f, row, "F열(cycle)")
        return {"type": "ID", "num": _num_str(num), "Cycle": int(cycle)}
    if b in ("Power", "Audio"):
        command = _require(d, row, "D열(command)")
        return {"type": b, "command": str(command)}
    if b == "delay":
        ms = _require(d, row, "D열(ms)")
        return {"type": "delay", "ms": int(ms)}
    if b in CAN_TYPES:
        message = _require(d, row, "D열(Message)")
        signal = _require(f, row, "F열(Signal)")
        value = _require(h, row, "H열(Value)")
        return {"type": b, "Message": str(message), "Signal": str(signal), "Value": str(value)}
    raise ScriptError(f"{row}행: 알 수 없는 스텝 종류 '{b}'")


def convert(ws) -> list[dict]:
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == HEADER_MARKER:
            header_row = r
            break
    if header_row is None:
        raise ScriptError(f"헤더 행(B열 == '{HEADER_MARKER}')을 찾을 수 없습니다 -- 시트 구조를 확인하세요")

    # stack[0] is the top-level step list; each "loop" row pushes a new frame
    # that a matching "]" row pops back into a nested {"type": "loop", ...} step.
    stack: list[dict] = [{"cycle": None, "start_row": None, "steps": []}]

    # Set while between a CANMsgS row and its matching CANMsgE row; collects
    # the member CANReq/CANEv rows into one merged Signals step (see the
    # module docstring). None outside a group.
    msg_group: Optional[dict] = None

    for r in range(header_row + 1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b is None:
            continue  # blank separator row

        if b == MSG_GROUP_START:
            if msg_group is not None:
                raise ScriptError(f"{r}행: 'CANMsgS'가 중첩되었습니다 (이미 {msg_group['start_row']}행에서 시작됨)")
            msg_group = {"type": None, "message": None, "signals": [], "start_row": r}
            continue

        if b == MSG_GROUP_END:
            if msg_group is None:
                raise ScriptError(f"{r}행: 대응하는 'CANMsgS' 시작 없이 'CANMsgE'가 나타났습니다")
            if not msg_group["signals"]:
                raise ScriptError(f"{msg_group['start_row']}행: 'CANMsgS'~'CANMsgE' 사이에 CAN 신호가 없습니다")
            stack[-1]["steps"].append(
                {"type": msg_group["type"], "Message": msg_group["message"], "Signals": msg_group["signals"]}
            )
            msg_group = None
            continue

        if msg_group is not None:
            if b not in GROUPABLE_CAN_TYPES:
                raise ScriptError(
                    f"{r}행: 'CANMsgS'~'CANMsgE' 안에는 CANReq/CANEv만 넣을 수 있습니다 (받은 값: '{b}')"
                )
            message = _require(ws.cell(row=r, column=4).value, r, "D열(Message)")
            signal = _require(ws.cell(row=r, column=6).value, r, "F열(Signal)")
            value = _require(ws.cell(row=r, column=8).value, r, "H열(Value)")
            if msg_group["message"] is None:
                msg_group["type"] = b
                msg_group["message"] = str(message)
            elif b != msg_group["type"]:
                raise ScriptError(
                    f"{r}행: 그룹 내 스텝 종류가 일치하지 않습니다 "
                    f"('{msg_group['start_row']}행 그룹은 '{msg_group['type']}', 이 행은 '{b}')"
                )
            elif str(message) != msg_group["message"]:
                raise ScriptError(
                    f"{r}행: 그룹 내 Message가 일치하지 않습니다 "
                    f"('{msg_group['start_row']}행 그룹은 '{msg_group['message']}', 이 행은 '{message}')"
                )
            msg_group["signals"].append({"Signal": str(signal), "Value": str(value)})
            continue

        if b == LOOP_END:
            if len(stack) == 1:
                raise ScriptError(f"{r}행: 대응하는 'loop' 시작 없이 ']'가 나타났습니다")
            frame = stack.pop()
            stack[-1]["steps"].append({"type": "loop", "cycle": frame["cycle"], "steps": frame["steps"]})
            continue

        if b == "loop":
            cycle = _require(ws.cell(row=r, column=4).value, r, "D열(cycle)")
            stack.append({"cycle": int(cycle), "start_row": r, "steps": []})
            continue

        if b not in STEP_TYPES:
            raise ScriptError(f"{r}행: 알 수 없는 스텝 종류 '{b}'")

        d = ws.cell(row=r, column=4).value
        f = ws.cell(row=r, column=6).value
        h = ws.cell(row=r, column=8).value
        stack[-1]["steps"].append(_build_step(r, b, d, f, h))

    if msg_group is not None:
        raise ScriptError(f"{msg_group['start_row']}행: 닫히지 않은 'CANMsgS'가 있습니다")
    if len(stack) != 1:
        unclosed = ", ".join(f"{frame['start_row']}행" for frame in stack[1:])
        raise ScriptError(f"닫히지 않은 loop가 있습니다 (시작 행: {unclosed})")
    return stack[0]["steps"]


def _count_leaf_steps(steps) -> int:
    n = 0
    for s in steps:
        n += _count_leaf_steps(s.children) if s.type == "loop" else 1
    return n


def _self_validate(steps: list[dict]) -> None:
    """Parse the converted JSON with the actual backend parser, so a
    structural mistake is caught here instead of surfacing later in the
    running app."""
    # This module now lives in backend/, so the parser is a sibling import.
    backend_dir = Path(__file__).resolve().parent
    sys.path.insert(0, str(backend_dir))
    try:
        from test_runner_service import parse_script
    except ImportError as exc:
        print(f"(참고: 백엔드 파서를 불러올 수 없어 자체 검증을 건너뜁니다 -- {exc})")
        return

    cases = parse_script(steps)
    total = sum(_count_leaf_steps(c.steps) for c in cases)
    print(f"자체 검증: {len(cases)}개 케이스, 총 {total}개 리프 스텝으로 파싱됨")
    for c in cases:
        print(f"  - case {c.num} (Cycle={c.cycle}): {_count_leaf_steps(c.steps)}개 스텝")


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", type=Path, help="입력 xlsx 파일 경로")
    parser.add_argument("-o", "--output", type=Path, help="출력 JSON 파일 경로 (기본: 입력 파일명.json)")
    parser.add_argument("--sheet", default="Script", help="변환할 시트 이름 (기본: Script)")
    args = parser.parse_args(argv)

    wb = openpyxl.load_workbook(args.input, data_only=True)
    if args.sheet not in wb.sheetnames:
        parser.error(f"시트 '{args.sheet}'를 찾을 수 없습니다 (시트 목록: {', '.join(wb.sheetnames)})")

    steps = convert(wb[args.sheet])

    output = args.output or args.input.with_suffix(".json")
    output.write_text(json.dumps(steps, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"{len(steps)}개 최상위 스텝 -> {output}")

    _self_validate(steps)


if __name__ == "__main__":
    try:
        main()
    except ScriptError as exc:
        print(f"변환 실패: {exc}", file=sys.stderr)
        sys.exit(1)