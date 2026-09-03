"""CAN evaluation environment backend.

FastAPI server running on the local PC. The browser GUI talks to this server
via REST (configuration) and WebSocket (real-time CAN RX stream / status),
and the server drives the USB-CAN adapter (PCAN / Vector CANcase) or an
in-process virtual bus.

Run:  uvicorn main:app --host 127.0.0.1 --port 8000
"""

import asyncio
import json
import logging
import os
import signal
import sys
import threading
import time
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Callable, Optional

if sys.platform == "win32":
    # Windows Ctrl-C/SIGINT delivery to asyncio's default ProactorEventLoop
    # is a well-documented weak spot, especially when the loop is kept
    # continuously busy (this app's WS broadcast loop ticks every 30ms, plus
    # constant HTTP polling from open widgets) -- confirmed on this project:
    # no "Shutting down" log and no cansim.shutdown log ever appeared despite
    # pressing Ctrl-C multiple times (Requirement.md). WindowsSelectorEvent
    # LoopPolicy's select()-based loop cooperates with Python's signal
    # checking far more reliably. Must be set before uvicorn creates its
    # event loop; uvicorn imports this module before starting the loop, so
    # module-import time (here, before any other asyncio use) is early
    # enough. This app doesn't use asyncio subprocesses anywhere (grepped),
    # so the one real limitation of the selector loop on Windows doesn't
    # apply here.
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

from fastapi import FastAPI, HTTPException, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import can
import diag_log
import timer_util
import isotp_service
from audio_service import AudioService, generate_monitor_filename
from can_manager import CanManager
from dbc_service import DbcService
from log_service import LogService
from power_supply_service import PowerSupplyService
from replay_service import ReplayService
from seedkey_client import SeedKeyService
from syslog_service import SysLogService
from can_log_service import CanLogService
from test_runner_service import TestRunnerService
from tx_scheduler import TxScheduler
from uds_download_manager import MultiUdsDownloadManager
from ota_tester_download_manager import OtaTesterDownloadManager


class _SuppressNoisyAccessLog(logging.Filter):
    """The frontend polls /api/testrunner/status every 400ms while the app
    is open (see canStore.ts / TestRunnerBox.tsx), which floods the terminal
    with access-log lines that carry no diagnostic value. Drop just that
    path; every other request still logs normally.

    /api/audio/level is the same story: AudioMonitorWidget polls it every
    100ms continuously (not just while its own Start is active) so it can
    reflect a recording started elsewhere -- see AudioMonitorWidget.tsx.
    /api/audio/waveform is polled at a similar cadence per zoomable chart."""

    _SUPPRESSED_PATHS = ("/api/testrunner/status", "/api/audio/level", "/api/audio/waveform")

    def filter(self, record: logging.LogRecord) -> bool:
        path = record.args[2] if record.args and len(record.args) > 2 else ""
        return not any(path.startswith(p) for p in self._SUPPRESSED_PATHS)


logging.getLogger("uvicorn.access").addFilter(_SuppressNoisyAccessLog())
logger = logging.getLogger(__name__)

# Diagnostic logging for the Windows "오디오 지연 확인 위젯 Start 후 CAN 전송이
# 매우 느려지고, Stop이 10여초 지연되며, 전역 Stop이 Failed to fetch로 실패한다"
# 조사 (Requirement.md 참고). uvicorn's own logging setup (run via `uvicorn
# main:app`) never configures the root logger, so a plain
# logging.getLogger(__name__).info(...)/.warning(...) elsewhere in this app
# would otherwise only surface WARNING+ through Python's silent
# last-resort handler -- give every "cansim.*" logger (audio_service,
# tx_scheduler, and this module's own diagnostics below) an explicit handler
# so both INFO and WARNING entries are always visible in the server console,
# independent of uvicorn's config.
_diag_handler = logging.StreamHandler()
_diag_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(name)s: %(message)s"))
diag_logger = logging.getLogger("cansim")
diag_logger.setLevel(logging.INFO)
diag_logger.addHandler(_diag_handler)
diag_logger.propagate = False

http_diag_logger = logging.getLogger("cansim.http")
_SLOW_REQUEST_MS = 300.0
_SLOW_BROADCAST_DRIFT_MS = 100.0
_inflight_requests = 0

# Ctrl-C still didn't stop the backend even after audio_service.shutdown()'s
# own 3s bound and uvicorn's --timeout-graceful-shutdown -- so the app's own
# lifespan shutdown block (below) had no visibility into which step it was
# actually stuck on. shutdown_logger.info() before/after each step gives that
# visibility, and _run_bounded() applies the same bounded-background-thread
# pattern already used in audio_service.shutdown() to the other two calls in
# that block that touch real hardware drivers with no timeout of their own
# (CanManager.disconnect()'s bus.shutdown(), PowerSupplyService.disconnect()'s
# VISA instrument close()) -- either is a plausible hang point on Windows if
# real hardware (not the virtual bus) is connected.
shutdown_logger = logging.getLogger("cansim.shutdown")
_SHUTDOWN_STEP_TIMEOUT_S = 3.0


def _run_bounded(fn: Callable[[], None], label: str, timeout_s: float = _SHUTDOWN_STEP_TIMEOUT_S) -> None:
    done = threading.Event()

    def _wrapped() -> None:
        try:
            fn()
        except Exception:
            shutdown_logger.warning("%s raised during shutdown", label, exc_info=True)
        finally:
            done.set()

    t0 = time.perf_counter()
    threading.Thread(target=_wrapped, daemon=True, name=f"shutdown-{label}").start()
    if done.wait(timeout=timeout_s):
        shutdown_logger.info("%s done in %.0fms", label, (time.perf_counter() - t0) * 1000.0)
    else:
        shutdown_logger.warning(
            "%s did not finish within %.0fs -- abandoning it so the process can still exit",
            label,
            timeout_s,
        )

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
LAYOUT_DIR = BASE_DIR / "layouts"
TESTRUNNER_LOG_DIR = BASE_DIR / "uploads" / "testrunner_logs"
TESTRUNNER_RESULT_DIR = BASE_DIR / "testrunner_results"
TESTRUNNER_AUDIO_DIR = BASE_DIR / "uploads" / "testrunner_audio"
TESTRUNNER_GOLDEN_DIR = BASE_DIR / "uploads" / "testrunner_golden"
CAN_LOG_DIR = BASE_DIR / "can_logs"
FRONTEND_DIST = BASE_DIR.parent / "frontend" / "dist"

can_manager = CanManager()
dbc_service = DbcService()
tx_scheduler = TxScheduler(can_manager, dbc_service)
replay_service = ReplayService(can_manager)
log_service = LogService(can_manager, CAN_LOG_DIR)
power_supply_service = PowerSupplyService()
audio_service = AudioService(TESTRUNNER_AUDIO_DIR, TESTRUNNER_GOLDEN_DIR)
syslog_service = SysLogService()
can_log_service = CanLogService(dbc_service)
test_runner_service = TestRunnerService(
    can_manager,
    dbc_service,
    tx_scheduler,
    replay_service,
    TESTRUNNER_LOG_DIR,
    TESTRUNNER_RESULT_DIR,
    power_service=power_supply_service,
    audio_service=audio_service,
)

seedkey_service = SeedKeyService()

uds_download_manager = MultiUdsDownloadManager(
    can_manager,
    isotp_service.send,
    isotp_service.receive,
    seedkey_service,
    log_dir=CAN_LOG_DIR,
)

ota_tester_manager = OtaTesterDownloadManager(
    can_manager,
    isotp_service.send,
    isotp_service.receive,
    seedkey_service,
    log_dir=CAN_LOG_DIR,
)

settings = {"ws_flush_ms": 30}
# global run gate: when stopped, no TX at all and the RX stream is discarded
run_state = {"running": True}
ws_clients: set[WebSocket] = set()


@asynccontextmanager
async def lifespan(app: FastAPI):
    UPLOAD_DIR.mkdir(exist_ok=True)
    LAYOUT_DIR.mkdir(exist_ok=True)
    TESTRUNNER_LOG_DIR.mkdir(parents=True, exist_ok=True)
    TESTRUNNER_RESULT_DIR.mkdir(exist_ok=True)
    TESTRUNNER_AUDIO_DIR.mkdir(parents=True, exist_ok=True)
    TESTRUNNER_GOLDEN_DIR.mkdir(parents=True, exist_ok=True)
    timer_util.enable_1ms_timer()
    broadcaster = asyncio.create_task(_broadcast_loop())
    yield
    shutdown_logger.info("lifespan shutdown starting")
    broadcaster.cancel()
    shutdown_logger.info("test_runner_service.stop() starting")
    test_runner_service.stop()
    shutdown_logger.info("replay_service.stop() starting")
    replay_service.stop()
    shutdown_logger.info("log_service.stop() starting")
    log_service.stop()
    shutdown_logger.info("tx_scheduler.shutdown() starting")
    tx_scheduler.shutdown()
    # bounded: bus.shutdown()/instrument.close() are real hardware-driver
    # calls with no timeout of their own -- see _run_bounded()'s docstring.
    _run_bounded(can_manager.disconnect, "can_manager.disconnect()")
    _run_bounded(power_supply_service.disconnect, "power_supply_service.disconnect()")
    # Previously missing entirely -- a stream left open (Start/Record never
    # Stopped) just stayed open through shutdown. audio_service.shutdown()
    # is itself bounded-time (see its docstring), so it can't turn into the
    # very "Ctrl-C doesn't exit" symptom this is meant to prevent.
    audio_service.shutdown()
    timer_util.disable_1ms_timer()
    shutdown_logger.info("lifespan shutdown complete")


app = FastAPI(title="CAN Evaluation Backend", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    # Wildcard would let any webpage the operator's browser has open make
    # cross-origin requests to this server (CAN TX, firmware upload, etc. --
    # a "drive-by localhost" attack). The only legitimate cross-origin caller
    # is the Vite dev server (see .claude/launch.json's --port --strictPort);
    # the production build is served from this same app (see FRONTEND_DIST
    # mount below), so it never needs CORS at all.
    allow_origins=["http://localhost:5174", "http://127.0.0.1:5174"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def _diag_timing_middleware(request, call_next):
    """Diagnostic-only: logs any request that takes longer than
    _SLOW_REQUEST_MS, with the in-flight request count at that moment.
    Sync `def` endpoints (every audio/tx/run endpoint in this file) run on
    Starlette's shared thread pool -- if that pool is ever saturated by piled-
    up audio polling requests, a manual CAN send (/api/tx/signal) or the
    global Stop (/api/run/stop) queues behind them and looks "stuck" from the
    browser's side. This surfaces exactly that queuing in the server log
    instead of it being invisible. Global counter is safe unaudited here: only
    the single asyncio event-loop thread ever touches it (this coroutine
    always runs there, even though call_next may await a threadpool-run sync
    handler)."""
    global _inflight_requests
    _inflight_requests += 1
    start = time.perf_counter()
    try:
        response = await call_next(request)
    finally:
        _inflight_requests -= 1
    dur_ms = (time.perf_counter() - start) * 1000.0
    if dur_ms > _SLOW_REQUEST_MS:
        # Rate-limited per path (not globally) -- otherwise a busy audio
        # endpoint throttling itself could hide /api/tx/signal or
        # /api/run/stop *also* going slow behind it, which is the exact
        # thing this log is meant to catch.
        n = diag_log.should_log(f"http.slow.{request.url.path}")
        if n >= 0:
            http_diag_logger.warning(
                "slow request: %s %s took %.0fms (in-flight now=%d)%s",
                request.method,
                request.url.path,
                dur_ms,
                _inflight_requests,
                diag_log.suffix(n),
            )
    return response


# ---- WebSocket: RX stream + status ------------------------------------


def _frame_to_dict(msg, do_decode: bool = True) -> dict:
    d = {
        "ts": msg.timestamp,
        "id": msg.arbitration_id,
        "ext": msg.is_extended_id,
        "dlc": msg.dlc,
        "data": msg.data.hex(),
        "fd": msg.is_fd,
        "brs": msg.bitrate_switch,
    }
    if do_decode and dbc_service.loaded:
        decoded = dbc_service.decode(msg.arbitration_id, bytes(msg.data))
        if decoded:
            d["decoded"] = decoded
    return d


async def _broadcast(payload: dict) -> None:
    if not ws_clients:
        return
    text = json.dumps(payload)
    # parallel send to avoid head-of-line blocking on slow clients
    results = await asyncio.gather(
        *[ws.send_text(text) for ws in list(ws_clients)], return_exceptions=True
    )
    dead = [ws for ws, r in zip(list(ws_clients), results) if isinstance(r, Exception)]
    for ws in dead:
        ws_clients.discard(ws)


async def _broadcast_loop() -> None:
    last_status = 0.0
    last_tick = time.monotonic()
    while True:
        await asyncio.sleep(settings["ws_flush_ms"] / 1000.0)
        now_tick = time.monotonic()
        drift_ms = (now_tick - last_tick) * 1000.0 - settings["ws_flush_ms"]
        if drift_ms > _SLOW_BROADCAST_DRIFT_MS:
            n = diag_log.should_log("http.broadcast_drift")
            if n >= 0:
                http_diag_logger.warning(
                    "broadcast loop tick delayed %.0fms beyond target %dms -- event loop "
                    "starved (GIL held elsewhere, e.g. audio_service numpy work in a "
                    "threadpool worker)%s",
                    drift_ms,
                    settings["ws_flush_ms"],
                    diag_log.suffix(n),
                )
        last_tick = now_tick
        frames = can_manager.drain_rx()
        if not run_state["running"]:
            frames = []  # discard RX while globally stopped
        if frames and ws_clients:
            # decode throttling: at 600µs burst, decoding 2000 frames stalls loop
            # only decode first 500 when overloaded, rest as raw
            do_decode_limit = 500
            await _broadcast(
                {"type": "rx", "frames": [_frame_to_dict(m, i < do_decode_limit) for i, m in enumerate(frames)]}
            )
        now = time.monotonic()
        if now - last_status >= 0.5:
            last_status = now
            try:
                status = _status()
            except Exception:
                logger.exception("_status() failed in broadcast loop")
            else:
                await _broadcast({"type": "status", **status})


@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket):
    await ws.accept()
    ws_clients.add(ws)
    try:
        await ws.send_text(json.dumps({"type": "status", **_status()}))
        while True:
            await ws.receive_text()  # keepalive; commands go through REST
    except WebSocketDisconnect:
        pass
    finally:
        ws_clients.discard(ws)


# ---- status / connection ----------------------------------------------


def _status() -> dict:
    # WS broadcast: keep payload small — only tail events for SWDL/OTA
    return {
        "can": can_manager.status(),
        "tx": tx_scheduler.status(),
        "replay": replay_service.info(),
        "dbc": {"loaded": dbc_service.loaded, "filename": dbc_service.filename},
        "settings": dict(settings),
        "run": dict(run_state),
        # lightweight summary only -- the full step-by-step event log is
        # fetched on demand via GET /api/testrunner/status, not broadcast
        # to every client every 0.5s.
        "test_runner": test_runner_service.summary(),
        "uds": uds_download_manager.all_status(tail=50),
        "ota_tester": ota_tester_manager.status(tail=100),
        "power": power_supply_service.info(),
        "audio": audio_service.info(),
        "log": log_service.status(),
    }


def _require_running() -> None:
    if not run_state["running"]:
        raise HTTPException(
            status_code=400, detail="전체 송수신이 정지 상태입니다 (Start를 누르세요)"
        )


@app.get("/api/status")
def get_status():
    return _status()


@app.post("/api/run/start")
def run_start():
    # "Start" means a clean restart, not "resume whatever was armed before" --
    # auto-periodic senders left over from signals touched before the last
    # Stop must not silently resume without the user touching that widget again.
    tx_scheduler.stop_auto()
    run_state["running"] = True
    tx_scheduler.set_paused(False)
    return _status()


@app.post("/api/run/stop")
def run_stop():
    run_state["running"] = False
    tx_scheduler.set_paused(True)
    tx_scheduler.stop_auto()
    replay_service.stop()
    test_runner_service.stop()
    # A running CAN-SWDL/OTA Tester download previously kept transmitting
    # in the background after the global Stop, since neither manager was
    # ever included here -- only each widget's own Stop button touched
    # them. Only stop slots that are actually mid-run (mgr.stop() on an
    # idle/READY slot would otherwise force its state back to IDLE for no
    # reason, per UdsDownloadManager.stop()'s own state-reset logic).
    for slot_index in range(MultiUdsDownloadManager.NUM_SLOTS):
        mgr = uds_download_manager.get_manager(slot_index)
        if mgr.running:
            mgr.stop()
    if ota_tester_manager.running:
        ota_tester_manager.stop()
    return _status()


class ConnectRequest(BaseModel):
    interface: str
    channel: str
    bitrate: int = 500000
    receive_own_messages: bool = True
    fd: bool = False
    data_bitrate: int = 2_000_000  # CAN-FD data-phase bitrate; ignored unless fd=True


@app.post("/api/connect")
def connect(req: ConnectRequest):
    # a fresh connect tears down the old notifier internally (see
    # CanManager.connect -> self.disconnect()), which would silently orphan
    # an in-progress recording's file handle -- flush and close it first.
    log_service.stop()
    try:
        return can_manager.connect(
            req.interface,
            req.channel,
            req.bitrate,
            req.receive_own_messages,
            fd=req.fd,
            data_bitrate=req.data_bitrate,
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/disconnect")
def disconnect():
    test_runner_service.stop()
    replay_service.stop()
    log_service.stop()
    tx_scheduler.stop()
    tx_scheduler.stop_auto()
    can_manager.disconnect()
    return _status()


@app.post("/api/log/start")
def log_start():
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    try:
        return log_service.start()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/log/stop")
def log_stop():
    return log_service.stop()


class SettingsRequest(BaseModel):
    ws_flush_ms: int


@app.post("/api/settings")
def update_settings(req: SettingsRequest):
    settings["ws_flush_ms"] = max(10, min(500, req.ws_flush_ms))
    return dict(settings)


# ---- DBC ----------------------------------------------------------------


@app.post("/api/dbc/upload")
async def upload_dbc(file: UploadFile):
    raw = await file.read()
    for encoding in ("utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    try:
        return dbc_service.load_string(text, file.filename or "uploaded.dbc")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"DBC parse error: {exc}")


@app.get("/api/dbc")
def get_dbc():
    return dbc_service.summary()


@app.get("/api/dbc/raw")
def get_dbc_raw():
    return dbc_service.raw() or {"loaded": False}


class SendTypeOverride(BaseModel):
    message_name: str
    signal_name: str
    send_type: str  # "event" | "periodic"


@app.post("/api/dbc/send-type")
def override_send_type(req: SendTypeOverride):
    try:
        dbc_service.set_send_type_override(
            req.message_name, req.signal_name, req.send_type
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return dbc_service.summary()


# ---- TX -----------------------------------------------------------------


class TxConfigRequest(BaseModel):
    entries: list[dict]


@app.post("/api/tx/configure")
def tx_configure(req: TxConfigRequest):
    try:
        return tx_scheduler.configure(req.entries)
    except (ValueError, KeyError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/tx/start")
def tx_start():
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    return tx_scheduler.start()


@app.post("/api/tx/stop")
def tx_stop():
    return tx_scheduler.stop()


class TxSendOnceRequest(BaseModel):
    arbitration_id: int
    data: str = ""
    is_extended: bool = False
    is_fd: bool = False
    bitrate_switch: bool = False
    key: Optional[str] = None


@app.post("/api/tx/send_once")
def tx_send_once(req: TxSendOnceRequest):
    """TX box row's Send button, and live edits to a row's data field while
    the list is running -- a one-shot send independent of that row's
    periodic flag."""
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    try:
        data = bytes.fromhex(req.data) if req.data else b""
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"잘못된 hex 데이터: {exc}")
    try:
        return tx_scheduler.send_once(
            req.arbitration_id, data, req.is_extended, req.is_fd, req.bitrate_switch, req.key,
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class SignalSendRequest(BaseModel):
    message_name: str
    values: dict[str, float | int | str]


@app.post("/api/tx/signal")
def tx_signal(req: SignalSendRequest):
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="no DBC loaded")
    try:
        return tx_scheduler.send_signal(req.message_name, req.values)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class EnablePeriodicRequest(BaseModel):
    rx_node: str = ""


@app.post("/api/tx/periodic/enable_all")
def tx_periodic_enable_all(req: EnablePeriodicRequest):
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="no DBC loaded")
    try:
        return tx_scheduler.enable_all_periodic(req.rx_node)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/tx/periodic/disable_all")
def tx_periodic_disable_all():
    return tx_scheduler.disable_all_periodic()


class AutoStopRequest(BaseModel):
    message_name: str | None = None


@app.post("/api/tx/auto/stop")
def tx_auto_stop(req: AutoStopRequest):
    return tx_scheduler.stop_auto(req.message_name)


class ValueGeneratorRequest(BaseModel):
    message_name: str
    signal_name: str
    mode: str  # "fixed" | "random" | "range"
    range_min: int | None = None
    range_max: int | None = None
    step: int = 1


@app.post("/api/tx/signal/generator")
def tx_signal_generator(req: ValueGeneratorRequest):
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="no DBC loaded")
    try:
        tx_scheduler.set_value_generator(
            req.message_name, req.signal_name, req.mode, req.range_min, req.range_max, req.step
        )
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return {"ok": True}


class GenerateSendRequest(BaseModel):
    message_name: str
    signal_name: str


@app.post("/api/tx/signal/generate")
def tx_signal_generate(req: GenerateSendRequest):
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="no DBC loaded")
    try:
        return tx_scheduler.send_generated(req.message_name, req.signal_name)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class InvalidSendRequest(BaseModel):
    message_name: str
    signal_name: str


@app.post("/api/tx/signal/invalid")
def tx_signal_invalid(req: InvalidSendRequest):
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="no DBC loaded")
    try:
        return tx_scheduler.send_invalid(req.message_name, req.signal_name)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class IsoTpSendRequest(BaseModel):
    tx_id: int
    fc_id: int
    data: str  # hex string, spaces allowed
    is_extended_id: bool = False
    fc_timeout_ms: int = 1000
    max_wait_frames: int = 10
    # Optional request/response mode: after sending, also wait for and
    # reassemble a reply on resp_id (used as isotp_service.receive()'s rx_id,
    # with this request's tx_id as the id we send our own Flow Control back
    # on) -- see Requirement.md's ISO-TP widget "응답 대기" entry. None (the
    # default) preserves the original send-only behavior.
    resp_id: Optional[int] = None
    resp_timeout_ms: int = 2000
    resp_fc_block_size: int = 0
    resp_fc_stmin: int = 0x00


@app.post("/api/isotp/send")
def isotp_send(req: IsoTpSendRequest):
    # blocking (waits for Flow Control, and optionally the response) -- runs
    # in FastAPI's threadpool since this handler is a plain `def`, so it does
    # not block the event loop.
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    hex_str = req.data.replace(" ", "").replace("\n", "").replace("\t", "")
    if len(hex_str) % 2 != 0:
        raise HTTPException(status_code=400, detail="데이터 hex 문자열의 길이가 홀수입니다")
    try:
        data = bytes.fromhex(hex_str)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"잘못된 hex 데이터: {exc}")
    if req.resp_id is None:
        try:
            return isotp_service.send(
                can_manager,
                req.tx_id,
                req.fc_id,
                data,
                is_extended_id=req.is_extended_id,
                fc_timeout_s=req.fc_timeout_ms / 1000.0,
                max_wait_frames=req.max_wait_frames,
            )
        except isotp_service.IsoTpError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    # "응답 대기" mode: registered *before* sending and shared between send()
    # (for any Flow Control it needs while sending) and receive(), so there
    # is no gap between them where python-can's Notifier has nowhere to
    # dispatch an arriving frame -- an ECU that answers unusually fast can
    # otherwise have its response land in that gap and be lost for good.
    # See isotp_service.send()/receive()'s own `reader` docstrings.
    if can_manager.notifier is None:
        raise HTTPException(status_code=400, detail="CAN 버스가 연결되어 있지 않습니다")
    reader = can.BufferedReader()
    can_manager.notifier.add_listener(reader)
    try:
        try:
            result = isotp_service.send(
                can_manager,
                req.tx_id,
                req.fc_id,
                data,
                is_extended_id=req.is_extended_id,
                fc_timeout_s=req.fc_timeout_ms / 1000.0,
                max_wait_frames=req.max_wait_frames,
                reader=reader,
            )
        except isotp_service.IsoTpError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

        try:
            response = isotp_service.receive(
                can_manager,
                req.resp_id,
                req.tx_id,
                timeout_s=req.resp_timeout_ms / 1000.0,
                is_extended_id=req.is_extended_id,
                fc_stmin=req.resp_fc_stmin,
                fc_block_size=req.resp_fc_block_size,
                reader=reader,
            )
            result["response"] = response.hex(" ").upper()
        except isotp_service.IsoTpError as exc:
            result["response_error"] = str(exc)
    finally:
        can_manager.notifier.remove_listener(reader)

    return result


# ---- Replay -------------------------------------------------------------


@app.post("/api/replay/upload")
async def upload_replay(file: UploadFile):
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".blf", ".asc"):
        raise HTTPException(status_code=400, detail="only .blf / .asc are supported")
    dest = UPLOAD_DIR / f"replay{suffix}"
    dest.write_bytes(await file.read())
    try:
        return replay_service.load(str(dest), file.filename)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"log parse error: {exc}")


class ReplayStartRequest(BaseModel):
    mode: str = "pass"  # "pass" | "stop"
    frame_ids: list[int] = []


@app.post("/api/replay/start")
def replay_start(req: ReplayStartRequest):
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    try:
        return replay_service.start(req.mode, req.frame_ids)
    except (ValueError, RuntimeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/replay/pause")
def replay_pause():
    try:
        return replay_service.pause()
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/replay/resume")
def replay_resume():
    try:
        return replay_service.resume()
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/replay/stop")
def replay_stop():
    return replay_service.stop()


# ---- Test scenario runner (Automation JSON scripts) -----------------------


@app.post("/api/testrunner/upload")
async def testrunner_upload_script(file: UploadFile):
    raw = await file.read()
    filename = file.filename or "script.json"
    suffix = Path(filename).suffix.lower()

    if suffix == ".xlsx":
        # Excel -> JSON 변환 (samples/xlsx_to_script.py 참고). 백엔드에서
        # openpyxl로 워크북을 읽고 xlsx_to_script.convert()로 steps를 만든 뒤
        # JSON 문자열로 직렬화하여 기존 test_runner_service.load() 파이프라인으로 전달.
        import io

        import openpyxl

        from xlsx_to_script import ScriptError, convert

        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Excel 파일을 열 수 없습니다: {exc}")
        sheet_name = "Script"
        if sheet_name not in wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"시트 '{sheet_name}'를 찾을 수 없습니다 (시트 목록: {', '.join(wb.sheetnames)})",
            )
        try:
            steps = convert(wb[sheet_name])
        except ScriptError as exc:
            raise HTTPException(status_code=400, detail=f"Excel 변환 오류: {exc}")
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Excel 파싱 오류: {exc}")
        text = json.dumps(steps, ensure_ascii=False, indent=2)
    else:
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("cp1252", errors="replace")

    try:
        return test_runner_service.load(text, filename)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"시나리오 파싱 오류: {exc}")


@app.get("/api/testrunner/script/raw")
def testrunner_script_raw():
    return test_runner_service.script_raw() or {"loaded": False}


def _safe_upload_filename(filename: Optional[str], default: str) -> str:
    """Basename only, with any path/traversal components stripped -- an
    UploadFile's filename is attacker-controlled input, and joining it
    directly onto an upload directory (as every handler below used to do)
    lets a crafted name like '../../etc/x' write outside that directory."""
    name = Path(filename or "").name
    return name or default


@app.post("/api/testrunner/logfile/upload")
async def testrunner_upload_logfile(file: UploadFile):
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".blf", ".asc"):
        raise HTTPException(status_code=400, detail="only .blf / .asc are supported")
    dest = TESTRUNNER_LOG_DIR / _safe_upload_filename(file.filename, f"log{suffix}")
    dest.write_bytes(await file.read())
    return {"saved": dest.name}


@app.post("/api/testrunner/start")
def testrunner_start():
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="no DBC loaded")
    try:
        return test_runner_service.start()
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/testrunner/pause")
def testrunner_pause():
    try:
        return test_runner_service.pause()
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/testrunner/resume")
def testrunner_resume():
    try:
        return test_runner_service.resume()
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/testrunner/stop")
def testrunner_stop():
    return test_runner_service.stop()


@app.get("/api/testrunner/status")
def testrunner_status():
    return test_runner_service.status()


@app.post("/api/testrunner/functions/upload")
async def testrunner_upload_functions(file: UploadFile):
    raw = await file.read()
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("cp1252", errors="replace")
    try:
        return test_runner_service.load_functions(text, file.filename or "functions.json")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"함수 마스터 JSON 파싱 오류: {exc}")


@app.get("/api/testrunner/functions/raw")
def testrunner_functions_raw():
    return test_runner_service.functions_raw() or {"loaded": False}


class FunctionStartRequest(BaseModel):
    name: str


@app.post("/api/testrunner/functions/start")
def testrunner_start_function(req: FunctionStartRequest):
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="no DBC loaded")
    try:
        return test_runner_service.start_function(req.name)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


# ---- UDS Software Download (CAN-SWDL) -----------------------------------


UDS_UPLOAD_DIR = BASE_DIR / "uploads" / "udswdl"


class UdsSwdlXmlUploadRequest(BaseModel):
    slot_index: int = 0


@app.post("/api/udswdl/xml/upload")
async def udswdl_xml_upload(file: UploadFile, slot_index: int = 0):
    """Load an XML procedure file into a slot (0/1/2)."""
    UDS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".xml":
        raise HTTPException(status_code=400, detail="only .xml files are supported")
    dest = UDS_UPLOAD_DIR / _safe_upload_filename(file.filename, "procedure.xml")
    content = await file.read()
    dest.write_bytes(content)
    try:
        manager = uds_download_manager.get_manager(slot_index)
        return {"slot": slot_index, "status": manager.load_xml(str(dest))}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"XML 파싱 오류: {exc}")


class UdsSwdlBinUploadRequest(BaseModel):
    slot_index: int = 0


@app.post("/api/udswdl/binary/upload")
async def udswdl_binary_upload(file: UploadFile, slot_index: int = 0):
    """Load a BIN file into a UDS slot (0/1/2)."""
    UDS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".bin":
        raise HTTPException(status_code=400, detail="only .bin files are supported")
    dest = UDS_UPLOAD_DIR / _safe_upload_filename(file.filename, "firmware.bin")
    content = await file.read()
    dest.write_bytes(content)
    try:
        manager = uds_download_manager.get_manager(slot_index)
        manager.load_binary(str(dest))
        return {"slot": slot_index, "status": manager.status()}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"BIN 파일 로드 오류: {exc}")


SEEDKEY_UPLOAD_DIR = BASE_DIR / "uploads" / "seedkey"


@app.post("/api/seedkey/upload")
async def seedkey_upload(file: UploadFile):
    """Load the real HKMC Advanced SeedKey DLL (Windows only) so
    SecurityAccess computes a real key instead of uds_core's dummy zero key."""
    SEEDKEY_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".dll":
        raise HTTPException(status_code=400, detail="only .dll files are supported")
    filename = _safe_upload_filename(file.filename, "seedkey.dll")
    dest = SEEDKEY_UPLOAD_DIR / filename
    content = await file.read()
    dest.write_bytes(content)
    try:
        return seedkey_service.load(str(dest), filename)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"SeedKey DLL 로드 오류: {exc}")


@app.get("/api/seedkey/status")
def seedkey_status():
    return seedkey_service.status()


class UdsSwdlStartRequest(BaseModel):
    slot_indices: list[int] = [0, 1, 2]
    # Single-array form (applies the same selection/overrides to every slot in
    # slot_indices). Kept for backward compat / single-slot callers.
    selected_steps: Optional[list[int]] = None
    modified_params: Optional[dict[str, dict[str, str]]] = None
    # Per-slot form (preferred when starting multiple slots at once): a map of
    # slot_index -> selection / params. Each per-slot entry takes precedence
    # over the shared fields above; a missing entry falls back to the shared
    # field, and a missing shared field means None (XML defaults / no overrides).
    per_slot_selected_steps: Optional[dict[int, Optional[list[int]]]] = None
    per_slot_modified_params: Optional[dict[int, Optional[dict[str, dict[str, str]]]]] = None
    global_stmin_tx: Optional[int] = None


@app.post("/api/udswdl/start")
def udswdl_start(req: UdsSwdlStartRequest):
    """Start UDS software download on specified slots (sequential).

    Each slot receives its own selected_steps / modified_params snapshot,
    threaded into that slot's worker thread as an immutable copy (see
    UdsDownloadManager.start) so concurrent slot starts cannot race.
    """
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    try:
        # Apply global stmin_tx override to all managers before starting
        if req.global_stmin_tx is not None:
            for idx in range(uds_download_manager.NUM_SLOTS):
                mgr = uds_download_manager.get_manager(idx)
                mgr._global_stmin_tx = req.global_stmin_tx
        return uds_download_manager.start_all(
            req.slot_indices,
            per_slot_selected_steps=req.per_slot_selected_steps,
            per_slot_modified_params=req.per_slot_modified_params,
            selected_steps=req.selected_steps,  # fallback when per-slot omitted
            modified_params=req.modified_params,  # fallback when per-slot omitted
        )
    except (RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/udswdl/stop")
def udswdl_stop(slot_index: int = 0):
    """Stop UDS download on a specific slot."""
    manager = uds_download_manager.get_manager(slot_index)
    return manager.stop()


class UdsSwdlParamRequest(BaseModel):
    slot_index: int = 0
    step_service: str
    params: dict[str, str]


@app.put("/api/udswdl/step_params")
def udswdl_set_params(req: UdsSwdlParamRequest):
    """Update parameters for a specific step on a specific slot."""
    manager = uds_download_manager.get_manager(req.slot_index)
    for key, value in req.params.items():
        manager.set_param(req.step_service, key, value)
    return manager.status()


@app.get("/api/udswdl/status")
def udswdl_status(tail: int | None = None):
    """Get UDS download status for all slots."""
    return uds_download_manager.all_status(tail=tail)


@app.get("/api/udswdl/steps")
def udswdl_steps(slot_index: int = 0):
    """Get parsed procedure steps for a specific slot."""
    manager = uds_download_manager.get_manager(slot_index)
    steps = manager.get_procedure_steps()
    if steps is None:
        raise HTTPException(status_code=400, detail="XML이 로드되지 않았습니다")
    return steps


# ---- Power supply (Phase 2) ------------------------------------------------


@app.post("/api/power/connect")
def power_connect():
    return power_supply_service.connect()


@app.post("/api/power/disconnect")
def power_disconnect():
    return power_supply_service.disconnect()


@app.get("/api/power/status")
def power_status():
    return power_supply_service.info()


# ---- 서버 종료 (Ctrl+C가 막히는 환경을 위한 대안 종료 경로) -----------------------
#
# Requirement.md 참고: PyVISA(NI-VISA 등 벤더 드라이버)가 파워서플라이 연결 시
# Windows 콘솔의 Ctrl+C(SIGINT) 전달 자체를 가로채는 것으로 보이는 사례가 실사용
# 중 확인됨(cansim.shutdown/cansim.power 로그가 전혀 안 찍히고 "Shutting down"도
# 안 뜸 -- 신호가 프로세스에 도달하기 전에 막힘). 이 엔드포인트는 콘솔 신호
# 경로를 전혀 타지 않는 일반 HTTP 요청이라 그 문제와 무관하게 항상 동작한다.


@app.post("/api/shutdown")
def shutdown_server():
    """가능한 범위에서 정리(lifespan의 종료 블록과 동일한 순서)한 뒤 프로세스를
    종료한다. 응답을 먼저 흘려보낼 시간을 준 뒤(0.3초) 백그라운드 스레드에서
    실행 -- 이 핸들러 자체가 즉시 프로세스를 죽이면 클라이언트가 응답을 못 받는다.
    `os.kill(getpid(), SIGTERM)`은 Windows에서 `TerminateProcess`로 처리돼(공식
    문서 기준 "프로세스를 무조건 종료") 콘솔 컨트롤 핸들러 경로를 전혀 타지 않으므로
    위 드라이버 문제와 무관하게 항상 프로세스를 끝낼 수 있다."""

    def _cleanup_and_exit() -> None:
        time.sleep(0.3)
        shutdown_logger.info("shutdown requested via /api/shutdown")
        try:
            tx_scheduler.shutdown()
        except Exception:
            shutdown_logger.warning("tx_scheduler.shutdown() failed", exc_info=True)
        _run_bounded(can_manager.disconnect, "can_manager.disconnect()")
        _run_bounded(power_supply_service.disconnect, "power_supply_service.disconnect()")
        try:
            audio_service.shutdown()
        except Exception:
            shutdown_logger.warning("audio_service.shutdown() failed", exc_info=True)
        shutdown_logger.info("terminating process now")
        os.kill(os.getpid(), signal.SIGTERM)

    threading.Thread(target=_cleanup_and_exit, daemon=True, name="api-shutdown").start()
    return {"ok": True, "message": "서버를 종료합니다"}


class PowerBatteryRequest(BaseModel):
    voltage: float
    current: float


@app.post("/api/power/battery")
def power_set_battery(req: PowerBatteryRequest):
    return power_supply_service.set_battery(req.voltage, req.current)


class PowerAccIgnRequest(BaseModel):
    command: str  # ACC_On | ACC_Off | IGN_On | IGN_Off | ACC_IGN_On | ACC_IGN_Off


@app.post("/api/power/acc_ign")
def power_set_acc_ign(req: PowerAccIgnRequest):
    return power_supply_service.set_acc_ign(req.command)


class PowerOnOffRepeatRequest(BaseModel):
    on_voltage: float
    on_current: float
    on_s: float
    off_voltage: float
    off_current: float
    off_s: float


@app.post("/api/power/onoff/start")
def power_onoff_start(req: PowerOnOffRepeatRequest):
    return power_supply_service.start_onoff_repeat(
        req.on_voltage, req.on_current, req.on_s, req.off_voltage, req.off_current, req.off_s
    )


@app.post("/api/power/onoff/stop")
def power_onoff_stop():
    return power_supply_service.stop_onoff_repeat()


class PowerSweepRequest(BaseModel):
    low: float
    high: float
    current: float
    leg_s: float


@app.post("/api/power/sweep/start")
def power_sweep_start(req: PowerSweepRequest):
    return power_supply_service.start_sweep(req.low, req.high, req.current, req.leg_s)


@app.post("/api/power/sweep/stop")
def power_sweep_stop():
    return power_supply_service.stop_sweep()


# ---- Audio (Phase 2) --------------------------------------------------------


@app.get("/api/audio/devices")
def audio_devices():
    return audio_service.refresh_devices()


class AudioDeviceRequest(BaseModel):
    index: int


@app.post("/api/audio/device")
def audio_select_device(req: AudioDeviceRequest):
    return audio_service.select_device(req.index)


@app.get("/api/audio/status")
def audio_status():
    return audio_service.info()


@app.post("/api/audio/monitor/start")
def audio_monitor_start():
    """오디오 신호 모니터 위젯의 Start -- 이미 테스트 러너 녹음이 진행 중이면
    새 스트림을 열지 않고 그 스트림의 레벨 데이터를 그대로 사용한다."""
    return audio_service.start_monitor()


@app.post("/api/audio/monitor/stop")
def audio_monitor_stop():
    """녹음이 소유한 스트림이면 끄지 않고 그대로 둔다 (모니터의 Stop이 실수로
    진행 중인 테스트 녹음을 끊지 않도록)."""
    return audio_service.stop_monitor()


@app.get("/api/audio/level")
def audio_level():
    return audio_service.get_level()


@app.get("/api/audio/waveform")
def audio_waveform(from_ms: float, to_ms: float, max_points: int = 300):
    """오디오 신호 모니터의 확대/축소 가능한 파형 차트가 폴링하는 엔드포인트.
    from_ms/to_ms는 JS Date.now() 단위 (epoch ms) -- 같은 로컬 머신이라 프론트/
    백엔드 시계가 그대로 맞아떨어진다."""
    return audio_service.get_waveform(from_ms, to_ms, max_points)


@app.post("/api/audio/record/start")
def audio_record_start():
    """오디오 신호 모니터 위젯의 Record 버튼 -- 파형을 보여주는 동시에 WAV로
    저장한다. 파일명은 타임스탬프로 자동 생성되고, 30분마다 새 파일로 이어서
    저장된다 (audio_service.py의 백그라운드 로테이션 타이머 참고)."""
    return audio_service.start_widget_recording(generate_monitor_filename())


@app.post("/api/audio/record/stop")
def audio_record_stop():
    return audio_service.stop_widget_recording()


@app.post("/api/audio/recording/stop")
def audio_recording_stop():
    """테스트 Sequence 실행기에서 러너가 시작한 녹음(owner='recording')을 위젯에서 Stop."""
    return audio_service.stop()


@app.post("/api/testrunner/golden/upload")
async def testrunner_upload_golden(file: UploadFile):
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".wav":
        raise HTTPException(status_code=400, detail="only .wav is supported")
    dest = TESTRUNNER_GOLDEN_DIR / _safe_upload_filename(file.filename, "golden.wav")
    dest.write_bytes(await file.read())
    return {"saved": dest.name}


# ---- sysLog 분석 ----------------------------------------------------------


@app.post("/api/syslog/upload")
async def syslog_upload_log(file: UploadFile):
    data = await file.read()
    try:
        return syslog_service.load_log(data, file.filename or "syslog.bin")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"sysLog parse error: {exc}")


@app.post("/api/syslog/db/upload")
async def syslog_upload_db(file: UploadFile):
    raw = await file.read()
    for encoding in ("utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    try:
        return syslog_service.load_db(text, file.filename or "sysLogDB.txt")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"sysLogDB parse error: {exc}")


@app.get("/api/syslog/status")
def syslog_status():
    return syslog_service.status()


@app.get("/api/syslog/timeline")
def syslog_timeline():
    return syslog_service.timeline()


@app.get("/api/syslog/ids")
def syslog_ids(segments: str | None = None):
    # segments 파라미터 자체가 없으면(None) 필터 없이 전체 개수. 파라미터가
    # 있으면(빈 문자열 포함) 그 인덱스들로 필터링 -- 빈 문자열은 "체크된 구간
    # 없음"(모든 ID count=0)을 뜻하므로, 파라미터 부재(필터 없음)와 구분해야
    # 한다.
    checked_segments = None
    if segments is not None:
        try:
            checked_segments = [int(x) for x in segments.split(",") if x.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="segments must be a comma-separated list of integers")
    return syslog_service.list_ids(checked_segments)


@app.get("/api/syslog/series")
def syslog_series(ids: str = ""):
    try:
        id_list = [int(x) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="ids must be a comma-separated list of integers")
    return syslog_service.get_series(id_list)


class SysLogScriptRequest(BaseModel):
    checked_segments: list[int]


@app.post("/api/syslog/generate_script")
def syslog_generate_script(req: SysLogScriptRequest):
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="DBC가 로드되지 않았습니다. 먼저 DBC를 업로드하세요.")
    dbc_summary = dbc_service.summary()
    return syslog_service.generate_test_script(
        req.checked_segments, dbc_summary["messages"], dbc_service.signal_send_type
    )


# ---- CAN log 분석 ---------------------------------------------------------

@app.post("/api/canlog/upload")
async def canlog_upload(file: UploadFile):
    data = await file.read()
    try:
        return can_log_service.load_log(data, file.filename or "canlog.blf")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"CAN log parse error: {exc}")


@app.get("/api/canlog/status")
def canlog_status():
    return can_log_service.status()


@app.get("/api/canlog/timeline")
def canlog_timeline():
    return can_log_service.timeline()


@app.get("/api/canlog/signals")
def canlog_signals():
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="DBC가 로드되지 않았습니다. 먼저 DBC를 업로드하세요.")
    return can_log_service.list_signals()


@app.get("/api/canlog/messages")
def canlog_messages():
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="DBC가 로드되지 않았습니다. 먼저 DBC를 업로드하세요.")
    return can_log_service.list_messages()


@app.get("/api/canlog/series")
def canlog_series(keys: str = ""):
    try:
        key_list = [k.strip() for k in keys.split(",") if k.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="keys must be comma-separated Message.Signal")
    return can_log_service.get_series(key_list)


class CanLogScriptRequest(BaseModel):
    range: Optional[dict] = None  # {"a_ms": int, "b_ms": int} | None
    rx_node: Optional[str] = None  # 필수 설정에서 선택한 RX 노드 (예: AMP_FD / AMP_HS)


@app.post("/api/canlog/generate_script")
def canlog_generate_script(req: CanLogScriptRequest):
    if not dbc_service.loaded:
        raise HTTPException(status_code=400, detail="DBC가 로드되지 않았습니다. 먼저 DBC를 업로드하세요.")
    if not req.rx_node:
        raise HTTPException(status_code=400, detail="RX 노드가 선택되지 않았습니다. 필수 설정에서 RX 노드(예: AMP_FD / AMP_HS)를 선택하세요.")
    if req.range is not None:
        try:
            a = int(req.range.get("a_ms", 0))
            b = int(req.range.get("b_ms", 0))
        except Exception:
            raise HTTPException(status_code=400, detail="range must contain a_ms and b_ms integers")
        if a == b:
            raise HTTPException(status_code=400, detail="커서 구간을 설정하세요 (A와 B가 같은 위치입니다)")
    dbc_summary = dbc_service.summary()
    return can_log_service.generate_test_script(
        req.range, dbc_summary["messages"], dbc_service.signal_send_type,
        rx_node=req.rx_node, dbc_nodes=dbc_summary.get("nodes"),
    )


# ---- Layout persistence --------------------------------------------------


def _layout_path(name: str) -> Path:
    safe = "".join(c for c in name if c.isalnum() or c in "-_ ").strip()
    if not safe:
        raise HTTPException(status_code=400, detail="invalid layout name")
    return LAYOUT_DIR / f"{safe}.json"


@app.get("/api/layouts")
def list_layouts():
    return {"layouts": sorted(p.stem for p in LAYOUT_DIR.glob("*.json"))}


@app.get("/api/layouts/{name}")
def get_layout(name: str):
    path = _layout_path(name)
    if not path.exists():
        raise HTTPException(status_code=404, detail="layout not found")
    return json.loads(path.read_text(encoding="utf-8"))


@app.post("/api/layouts/{name}")
async def save_layout(name: str, body: dict):
    _layout_path(name).write_text(
        json.dumps(body, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return {"saved": name}


@app.delete("/api/layouts/{name}")
def delete_layout(name: str):
    path = _layout_path(name)
    if path.exists():
        path.unlink()
    return {"deleted": name}


# ---- OTA Tester (GITAuto test-rule XML) ----------------------------------


OTA_TESTER_UPLOAD_DIR = BASE_DIR / "uploads" / "ota_tester"


@app.get("/api/ota_tester/status")
def ota_tester_status(tail: int | None = None):
    """Get OTA Tester status."""
    return ota_tester_manager.status(tail=tail)


@app.post("/api/ota_tester/case/xml_upload")
async def ota_tester_case_xml_upload(
    file: UploadFile, case_id: str, label: str, kind: str = "testBlock",
    order: int = 0, enabled: bool = True,
):
    """Load one hook/testBlock's test-rule XML as a case in the run sequence.

    ``case_id`` is the hook/testBlock's own id (a stable key so re-uploading
    the same case, e.g. after re-selecting the folder, replaces it in place
    rather than appending a duplicate); ``order`` fixes its position in the
    sequence (hooks before testBlocks, then each list's own JSON order)."""
    OTA_TESTER_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".xml":
        raise HTTPException(status_code=400, detail="only .xml files are supported")
    dest = OTA_TESTER_UPLOAD_DIR / _safe_upload_filename(f"{case_id}_{file.filename}", "test_rule.xml")
    content = await file.read()
    dest.write_bytes(content)
    try:
        return ota_tester_manager.add_case(case_id, label, kind, str(dest), order, enabled)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"XML 파싱 오류: {exc}")


@app.post("/api/ota_tester/case/binary_upload")
async def ota_tester_case_binary_upload(file: UploadFile, case_id: str):
    """Load the firmware binary referenced by one case's binaryPath."""
    OTA_TESTER_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".bin":
        raise HTTPException(status_code=400, detail="only .bin files are supported")
    dest = OTA_TESTER_UPLOAD_DIR / _safe_upload_filename(f"{case_id}_{file.filename}", "firmware.bin")
    content = await file.read()
    dest.write_bytes(content)
    try:
        return ota_tester_manager.set_case_binary(case_id, str(dest))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"BIN 파일 로드 오류: {exc}")


class OtaTesterCaseEnableRequest(BaseModel):
    case_id: str
    enabled: bool


@app.post("/api/ota_tester/case/enable")
def ota_tester_case_enable(req: OtaTesterCaseEnableRequest):
    """Toggle whether a single case is included in the next run."""
    try:
        return ota_tester_manager.set_case_enabled(req.case_id, req.enabled)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class OtaTesterSetAllEnabledRequest(BaseModel):
    enabled: bool


@app.post("/api/ota_tester/cases/set_all_enabled")
def ota_tester_set_all_enabled(req: OtaTesterSetAllEnabledRequest):
    """전체 선택 / 전체 해제."""
    return ota_tester_manager.set_all_enabled(req.enabled)


@app.post("/api/ota_tester/cases/clear")
def ota_tester_cases_clear():
    """Reset the case list before loading a (possibly different) folder."""
    try:
        return ota_tester_manager.clear_cases()
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/api/ota_tester/case/steps")
def ota_tester_case_steps(case_id: str):
    """Per-command checklist data for one case (mirrors /api/udswdl/steps)."""
    try:
        return ota_tester_manager.get_case_steps(case_id)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class OtaTesterCaseSelectedStepsRequest(BaseModel):
    case_id: str
    selected_steps: Optional[list[int]] = None


@app.put("/api/ota_tester/case/selected_steps")
def ota_tester_case_selected_steps(req: OtaTesterCaseSelectedStepsRequest):
    """Set which step indices within a case run (None = all, [] = none)."""
    try:
        return ota_tester_manager.set_case_selected_steps(req.case_id, req.selected_steps)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


class OtaTesterStartRequest(BaseModel):
    request_id: int
    response_id: int
    global_stmin_tx: Optional[int] = None


@app.post("/api/ota_tester/start")
def ota_tester_start(req: OtaTesterStartRequest):
    """Start OTA Tester procedure."""
    _require_running()
    if not can_manager.connected:
        raise HTTPException(status_code=400, detail="CAN bus is not connected")
    try:
        if req.global_stmin_tx is not None:
            ota_tester_manager._global_stmin_tx = req.global_stmin_tx
        return ota_tester_manager.start(req.request_id, req.response_id)
    except (RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/ota_tester/stop")
def ota_tester_stop():
    """Stop OTA Tester procedure."""
    return ota_tester_manager.stop()


# ---- Frontend static files (production build) ----------------------------

if FRONTEND_DIST.exists():
    app.mount("/", StaticFiles(directory=FRONTEND_DIST, html=True), name="frontend")
