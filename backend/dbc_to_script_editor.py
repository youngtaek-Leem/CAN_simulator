#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
dbc_to_script_editor.py
========================
CAN DBC(.dbc) 파일을 읽어서, 아래 절차대로 테스트 스크립트를 쉽게 입력할 수 있는
엑셀 파일(.xlsx)을 생성합니다.

절차 (Script 시트)
-------------------
  B열에 스텝 종류 선택: ID / Power / delay / CANReq / Audio / CANEv / loop
    - ID     : num(D열) 자동 1씩 증가, F열에 cycle 숫자 직접 입력
    - Power  : command(C열) 자동표시, D열 드롭다운에서 명령어 선택
    - delay  : ms(C열) 자동표시, D열에 숫자 직접 입력
    - CANReq : F열에 CAN 신호명을 입력(일부만 입력해도 검색됨)/선택 -> Message(D열) 자동 채움, H열에 Value 직접 입력
    - Audio  : command(C열) 자동표시, D열 드롭다운에서 명령어 선택
    - CANEv  : CANReq와 동일
    - loop   : D열에 반복횟수 직접 입력, F열에 '[' 자동표시
  J열에는 CAN DB(VAL_)에 정의된 유효값이 참고용으로 자동 표시됩니다.

사용법
------
    python3 dbc_to_script_editor.py <입력.dbc> [출력.xlsx]

필요 패키지: openpyxl (pip install openpyxl)
"""

import argparse
import re
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.formatting.rule import FormulaRule


# ---------------------------------------------------------------------------
# 1. DBC 파싱
# ---------------------------------------------------------------------------

BO_RE = re.compile(r'^BO_ (\d+) (\S+): (\d+) (\S+)')
SG_RE = re.compile(r'^\s*SG_ (\w+)\s*(?:M|m\d+)?\s*:')
CYCLE_RE = re.compile(r'^BA_ "GenMsgCycleTime" BO_ (\d+) (\d+)\s*;')
CYCLE_DEFAULT_RE = re.compile(r'^BA_DEF_DEF_\s+"GenMsgCycleTime"\s+(\d+)\s*;')
VAL_RE = re.compile(r'^VAL_ (\d+) (\w+) (.+);\s*$')
PAIR_RE = re.compile(r'(-?\d+)\s+"((?:[^"\\]|\\.)*)"')


def parse_dbc_messages(text):
    """(signal, message, msg_id, cycle_ms) 리스트, 신호명 중복 딕셔너리를 반환."""
    messages = {}
    cur_id = None
    default_cycle = 0

    for line in text.splitlines():
        m = BO_RE.match(line)
        if m:
            cur_id = int(m.group(1))
            messages[cur_id] = {"name": m.group(2), "signals": [], "cycle": None}
            continue
        m = SG_RE.match(line)
        if m and cur_id is not None:
            messages[cur_id]["signals"].append(m.group(1))
            continue
        m = CYCLE_DEFAULT_RE.match(line)
        if m:
            default_cycle = int(m.group(1))

    for line in text.splitlines():
        m = CYCLE_RE.match(line)
        if m:
            mid = int(m.group(1))
            if mid in messages:
                messages[mid]["cycle"] = int(m.group(2))

    if not messages:
        raise ValueError("DBC 파일에서 메시지(BO_)를 하나도 찾지 못했습니다.")

    rows = []
    dup_check = defaultdict(list)
    for mid, info in messages.items():
        cycle = info["cycle"] if info["cycle"] is not None else default_cycle
        for sig in info["signals"]:
            rows.append((sig, info["name"], mid, cycle))
            dup_check[sig].append(info["name"])

    rows.sort(key=lambda r: r[0].upper())
    duplicates = {s: msgs for s, msgs in dup_check.items() if len(msgs) > 1}
    return rows, messages, duplicates


def parse_dbc_values(text):
    """(signal, value_dec, value_hex, description) 리스트를 반환 (VAL_ 라인)."""
    entries = []
    for line in text.splitlines():
        m = VAL_RE.match(line)
        if m:
            _mid, sig, rest = m.groups()
            for val_s, desc in PAIR_RE.findall(rest):
                v = int(val_s)
                hexs = f"0x{(v & 0xFFFFFFFF):02X}" if v < 0 else f"0x{v:02X}"
                desc = desc.replace("\r", " ").replace("\n", " ").strip()
                entries.append((sig, v, hexs, desc))
    return entries


def parse_dbc(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        text = f.read()
    sig_rows, messages, duplicates = parse_dbc_messages(text)
    val_entries = parse_dbc_values(text)
    return sig_rows, messages, duplicates, val_entries


# ---------------------------------------------------------------------------
# 2. 엑셀 생성
# ---------------------------------------------------------------------------

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TYPE_FILL = PatternFill("solid", fgColor="DDEBF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
AUTO_FILL = PatternFill("solid", fgColor="F2F2F2")
REF_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STEP_TYPES = ["ID", "CANReq", "CANEv", "CANResp", "CANMsgS", "CANMsgE", "delay", "Audio", "Power", "loop", "]"]
POWER_CMDS = ["ACC_IGN_ON", "ACC_IGN_OFF", "ACC_ON", "IGN_ON", "ACC_OFF", "IGN_OFF"]
AUDIO_CMDS = ["StartREC", "StopREC", "compWAV", "RMSmeasure"]
N_ROWS = 200  # Script 시트에 미리 준비해 둘 작업 행 수


def build_workbook(sig_rows, val_entries, dbc_filename, out_path):
    n_sig = len(sig_rows)
    n_val = len(val_entries)

    # 메시지별 신호 목록 (Message -> Signal 종속 드롭다운용)
    by_msg = defaultdict(list)
    for sig, msg, mid, cyc in sig_rows:
        by_msg[msg].append(sig)
    for msg in by_msg:
        by_msg[msg].sort(key=str.upper)
    all_messages = sorted(by_msg.keys(), key=str.upper)
    max_sig_per_msg = max(len(v) for v in by_msg.values())

    wb = Workbook()

    # -- Sheet: Script ---------------------------------------------------
    ws = wb.active
    ws.title = "Script"
    ws.sheet_view.showGridLines = False

    ws["B2"] = "CAN 테스트 스크립트 입력기"
    ws["B2"].font = Font(name=FONT, size=14, bold=True, color="1F4E78")

    instructions = (
        "① B열(연한 파랑)에 스텝 종류를 선택하세요: ID / Power / delay / CANReq / Audio / CANEv / loop\n"
        "② A/C/E/G열(회색)은 스텝 종류에 따라 라벨과 값이 자동으로 채워집니다 (직접 입력하지 마세요).\n"
        "③ D/F/H열(노란색)은 스텝 종류에 따라 직접 입력하거나 드롭다운에서 선택하는 칸입니다.\n"
        "    - ID: F열에 cycle 숫자를 입력하세요 (num은 D열에 자동으로 1씩 증가하며 채워집니다).\n"
        "    - Power/Audio: D열 드롭다운에서 명령어를 선택하세요.\n"
        "    - delay/loop: D열에 숫자(ms 또는 반복횟수)를 입력하세요.\n"
        "    - CANReq/CANEv: I열(하늘색)에 CAN 신호명의 일부를 입력하면 포함된 신호가 검색됩니다.\n"
        "      선택하면 F열(Signal)과 D열(Message)이 자동으로 채워집니다.\n"
        "      검색 없이 바로 입력하려면 F열 드롭다운에서 직접 선택하거나 타이핑하세요.\n"
        "      K열에는 DBC(VAL_)에 정의된 유효값이 참고용으로 자동 표시됩니다.\n"
        "④ J열(연두색)은 CAN DB(VAL_)에 정의된 유효값 목록을 참고용으로 자동 표시합니다.\n"
        "⑤ 드롭다운에 없는 값을 입력하면 '계속하시겠습니까?' 경고가 뜰 수 있습니다 — 오류가 아니며, 확인 후 진행하면 됩니다."
    )
    ws["B3"] = instructions
    ws["B3"].font = Font(name=FONT, size=10, color="404040")
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B3:K3")
    ws.row_dimensions[3].height = 140

    HEADER_ROW = 7
    col_labels = {
        "A": "(자동) type",          "B": "Step 종류 선택",
        "C": "라벨1 / Signal 검색",  "D": "(자동) Message",
        "E": "(자동) 라벨2",          "F": "(자동) Signal",
        "G": "(자동) 라벨3",          "H": "Value (입력/선택)",
        "I": "(자동) Value 설명",     "J": "DBC 참고값 (자동)",
    }
    for c, h in col_labels.items():
        cell = ws[f"{c}{HEADER_ROW}"]
        cell.value = h
        cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[HEADER_ROW].height = 28

    first_row = HEADER_ROW + 1
    last_row = HEADER_ROW + N_ROWS

    for i in range(N_ROWS):
        r = first_row + i

        a = ws[f"A{r}"]
        a.value = f'=IF($B{r}<>"","type","")'
        a.font = Font(name=FONT, size=10, color="808080")
        a.fill = AUTO_FILL
        a.border = BORDER
        a.alignment = Alignment(horizontal="center")

        b = ws[f"B{r}"]
        b.font = Font(name=FONT, size=10, bold=True)
        b.fill = TYPE_FILL
        b.border = BORDER

        # C: CANReq/CANEv → Signal 검색·선택 입력칸 (노란색)
        #    나머지 → 라벨 자동 표시 (회색)
        c = ws[f"C{r}"]
        is_can = f'OR($B{r}="CANReq",$B{r}="CANEv",$B{r}="CANResp")'
        c.value = (
            f'=IF({is_can},"",'
            f'IF($B{r}="ID","num",IF($B{r}="Power","command",'
            f'IF($B{r}="delay","ms",IF($B{r}="Audio","command",'
            f'IF($B{r}="loop","cycle",""))))))'
        )
        c.font = Font(name=FONT, size=10, color="808080")
        c.fill = AUTO_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

        # D: CANReq/CANEv/CANResp → Message 자동(C열 Signal 기준) / ID → num 자동 / 나머지 빈값
        d = ws[f"D{r}"]
        d.value = (
            f'=IF($B{r}="ID",COUNTIF($B${first_row}:$B{r},"ID"),'
            f'IF(OR($B{r}="CANReq",$B{r}="CANEv",$B{r}="CANResp"),'
            f'IFERROR(INDEX(DB!$B:$B,MATCH($C{r},DB!$A:$A,0)),""),""))'
        )
        d.font = Font(name=FONT, size=10, color="505050")
        d.fill = AUTO_FILL
        d.border = BORDER
        d.alignment = Alignment(horizontal="right")

        # E: Signal 라벨 / ID → cycle / loop → steps
        e = ws[f"E{r}"]
        e.value = (
            f'=IF($B{r}="ID","cycle",'
            f'IF(OR($B{r}="CANReq",$B{r}="CANEv",$B{r}="CANResp"),"Signal",'
            f'IF($B{r}="loop","steps","")))'
        )
        e.font = Font(name=FONT, size=10, color="808080")
        e.fill = AUTO_FILL
        e.border = BORDER
        e.alignment = Alignment(horizontal="center")

        # F: CANReq/CANEv/CANResp → C열 값 자동 복사 / loop → ']' / 나머지 빈값
        f = ws[f"F{r}"]
        f.value = (
            f'=IF(OR($B{r}="CANReq",$B{r}="CANEv",$B{r}="CANResp"),$C{r},'
            f'IF($B{r}="loop","]",""))'
        )
        f.font = Font(name=FONT, size=10, color="505050")
        f.fill = AUTO_FILL
        f.border = BORDER
        f.alignment = Alignment(horizontal="right")

        # G: Value 라벨
        g = ws[f"G{r}"]
        g.value = f'=IF(OR($B{r}="CANReq",$B{r}="CANEv",$B{r}="CANResp"),"Value","")'
        g.font = Font(name=FONT, size=10, color="808080")
        g.fill = AUTO_FILL
        g.border = BORDER
        g.alignment = Alignment(horizontal="center")

        # H: Value 입력/선택
        h = ws[f"H{r}"]
        h.font = Font(name=FONT, size=10)
        h.fill = INPUT_FILL
        h.border = BORDER
        h.alignment = Alignment(horizontal="right")

        # I: Value 설명 자동 (C열 Signal + H열 hex값 → DB_VAL 복합키 조회)
        i_cell = ws[f"I{r}"]
        i_cell.value = (
            f'=IFERROR(INDEX(DB_VAL!$D:$D,'
            f'MATCH($C{r}&"|"&$H{r},DB_VAL!$E:$E,0)),"")'
        )
        i_cell.font = Font(name=FONT, size=9, color="375623")
        i_cell.fill = REF_FILL
        i_cell.border = BORDER
        i_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # J: DBC 전체 참고값 (C열 Signal 기준)
        j = ws[f"J{r}"]
        formula_text = (
            f'=IFERROR(_xlfn.TEXTJOIN(" | ",TRUE,'
            f'IF(DB_VAL!$A$2:$A${n_val+1}=$C{r},'
            f'DB_VAL!$C$2:$C${n_val+1}&"="&DB_VAL!$D$2:$D${n_val+1},"")),"")'
        )
        j.value = ArrayFormula(ref=f"J{r}", text=formula_text)
        j.font = Font(name=FONT, size=9, color="375623")
        j.fill = REF_FILL
        j.border = BORDER
        j.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 28   # Signal 검색 입력칸
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 30   # Value 설명
    ws.column_dimensions["J"].width = 55   # DBC 참고값
    ws.freeze_panes = f"B{first_row}"

    # C열 조건부 서식: CANReq/CANEv/CANResp일 때 노란색(입력칸)으로 표시
    # FormulaRule에 fill을 직접 넣으면 dxf의 patternType이 누락되어 Excel 오류 발생 →
    # Rule + DifferentialStyle로 명시적 구성
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import Rule as _Rule
    dxf = DifferentialStyle(fill=PatternFill(patternType="solid", fgColor="FFF2CC"))
    ws.conditional_formatting.add(
        f"C{first_row}:C{last_row}",
        _Rule(
            type="expression",
            dxf=dxf,
            formula=[f'OR($B{first_row}="CANReq",$B{first_row}="CANEv",$B{first_row}="CANResp")'],
        )
    )

    # -- Sheet: DB ---------------------------------------------------------
    db = wb.create_sheet("DB")
    for c, h in zip(["A", "B", "C", "D"], ["CAN 신호명 (Signal)", "CAN Message", "Message ID (dec)", "Cycle Time [ms]"]):
        cell = db[f"{c}1"]
        cell.value = h
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for idx, (sig, msg, mid, cyc) in enumerate(sig_rows):
        r = idx + 2
        db[f"A{r}"] = sig
        db[f"B{r}"] = msg
        db[f"C{r}"] = mid
        db[f"D{r}"] = cyc
        for c in ["A", "B", "C", "D"]:
            db[f"{c}{r}"].font = Font(name=FONT, size=10)
    db.column_dimensions["A"].width = 32
    db.column_dimensions["B"].width = 26
    db.column_dimensions["C"].width = 16
    db.column_dimensions["D"].width = 14
    db.freeze_panes = "A2"
    tab = Table(displayName="tblCANDB", ref=f"A1:D{n_sig + 1}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
    db.add_table(tab)

    # -- Sheet: DB_VAL -------------------------------------------------------
    dbv = wb.create_sheet("DB_VAL")
    for c, h in zip(["A", "B", "C", "D", "E"],
                    ["CAN 신호명 (Signal)", "Value (dec)", "Value (hex)", "설명 (Description)", "키(Signal|hex)"]):
        cell = dbv[f"{c}1"]
        cell.value = h
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for idx, (sig, vdec, vhex, desc) in enumerate(val_entries):
        r = idx + 2
        dbv[f"A{r}"] = sig
        dbv[f"B{r}"] = vdec
        dbv[f"C{r}"] = vhex
        dbv[f"D{r}"] = desc
        dbv[f"E{r}"] = f"{sig}|{vhex}"   # 복합 조회 키 (Signal + hex값)
        for c in ["A", "B", "C", "D", "E"]:
            dbv[f"{c}{r}"].font = Font(name=FONT, size=9)
    dbv.column_dimensions["A"].width = 32
    dbv.column_dimensions["B"].width = 12
    dbv.column_dimensions["C"].width = 12
    dbv.column_dimensions["D"].width = 60
    dbv.column_dimensions["E"].width = 0   # 숨김 (너비 0)
    dbv.freeze_panes = "A2"
    tab2 = Table(displayName="tblCANVAL", ref=f"A1:E{n_val + 1}")
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
    dbv.add_table(tab2)

    # -- Sheet: Lists (hidden) -----------------------------------------------
    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    lists["A1"] = "StepType"
    for i, v in enumerate(STEP_TYPES):
        lists.cell(row=i + 2, column=1, value=v)
    lists["B1"] = "Power"
    for i, v in enumerate(POWER_CMDS):
        lists.cell(row=i + 2, column=2, value=v)
    lists["C1"] = "Audio"
    for i, v in enumerate(AUDIO_CMDS):
        lists.cell(row=i + 2, column=3, value=v)

    # -- Sheet: DB_VAL_hex (hidden) -------------------------------------------
    # Signal 하나당 한 열: 그 Signal에 속한 hex값 목록 (H열 Value 참고용)
    from collections import defaultdict as _dd
    val_by_sig = _dd(list)
    for sig, vdec, vhex, desc in val_entries:
        val_by_sig[sig].append(vhex)
    all_val_sigs = sorted(val_by_sig.keys(), key=str.upper)

    dbvh = wb.create_sheet("DB_VAL_hex")
    dbvh.sheet_state = "hidden"
    for ci, sig in enumerate(all_val_sigs):
        col = ci + 1
        dbvh.cell(row=1, column=col, value=sig)
        for ri, vhex in enumerate(val_by_sig[sig]):
            dbvh.cell(row=ri + 2, column=col, value=vhex)

    # -- 데이터 유효성 검사 --------------------------------------------------
    # 이름 정의(Named range)와 OFFSET 기반 동적 드롭다운을 모두 제거하고,
    # 시트 범위를 직접 참조하는 가장 단순한 형태로 구성한다.
    # (이름 정의가 있으면 일부 Excel 환경에서 파일이 손상되었다는 오류를
    #  일으키며 이름 정의 전체와 무관한 유효성 검사까지 함께 삭제되는
    #  문제가 확인되어, 안정성을 최우선으로 단순한 방식으로 전환함)
    dv_type = DataValidation(
        type="list",
        formula1=f"=Lists!$A$2:$A${1+len(STEP_TYPES)}",
        allow_blank=True, showDropDown=False,
    )
    dv_type.error = "목록에 있는 스텝 종류만 선택할 수 있습니다."
    dv_type.errorTitle = "잘못된 Step 종류"
    dv_type.errorStyle = "stop"
    dv_type.prompt = "ID / CANReq / CANEv / CANResp / CANMsgS / CANMsgE / delay / Audio / Power / loop / ] 중 선택하세요."
    dv_type.promptTitle = "Step 종류 선택"
    ws.add_data_validation(dv_type)
    dv_type.add(f"B{first_row}:B{last_row}")

    # C: Signal 검색 (DB 시트 범위 직접 참조 → Excel 실시간 검색 가능)
    dv_c = DataValidation(
        type="list",
        formula1=f"=DB!$A$2:$A${n_sig+1}",
        allow_blank=True, showDropDown=False,
    )
    dv_c.error = "DB에 없는 CAN 신호명입니다."
    dv_c.errorTitle = "확인 필요"
    dv_c.errorStyle = "warning"
    dv_c.prompt = "CANReq/CANEv/CANResp: 일부 문자만 입력해도 포함된 Signal이 검색됩니다. 선택하면 D열(Message)·F열(Signal)이 자동으로 채워집니다."
    dv_c.promptTitle = "Signal 검색"
    ws.add_data_validation(dv_c)
    dv_c.add(f"C{first_row}:C{last_row}")

    # D: Power/Audio 명령어 목록 (Lists 시트 범위를 문자열로 직접 지정)
    dv_d = DataValidation(
        type="list",
        formula1=(
            f'=INDIRECT(IF($B{first_row}="Power","Lists!$B$2:$B${1+len(POWER_CMDS)}",'
            f'IF($B{first_row}="Audio","Lists!$C$2:$C${1+len(AUDIO_CMDS)}","")))'
        ),
        allow_blank=True, showDropDown=False,
    )
    dv_d.error = "목록에 없는 값입니다. (숫자를 직접 입력하는 경우에는 무시하고 계속하세요)"
    dv_d.errorTitle = "확인 필요"
    dv_d.errorStyle = "warning"
    dv_d.prompt = "Power/Audio: 명령어 선택 / delay·loop·ID: 숫자 직접 입력"
    dv_d.promptTitle = "값 입력"
    ws.add_data_validation(dv_d)
    dv_d.add(f"D{first_row}:D{last_row}")

    # H: Value는 드롭다운 없이 직접 입력 (OFFSET 기반 동적 드롭다운이 일부
    # Excel 환경에서 파일 손상 오류를 일으켜 제거함). I열에 선택한 값의
    # 설명이 자동 표시되고, J열에 그 Signal의 전체 유효값 목록이 참고용으로
    # 표시되므로, 그것을 보고 H열에 값을 입력하면 된다.

    # -- Sheet: 사용법 -----------------------------------------------------
    help_ws = wb.create_sheet("사용법")
    help_ws.sheet_view.showGridLines = False
    help_ws.column_dimensions["B"].width = 100

    def add_line(r, text, size=10, bold=False, color=None):
        cell = help_ws[f"B{r}"]
        cell.value = text
        cell.font = Font(name=FONT, size=size, bold=bold, color=color or ("1F4E78" if bold else "333333"))
        return r + 1

    r = 2
    r = add_line(r, "이 파일 사용법", 13, True)
    r += 1
    r = add_line(r, "[스텝 종류별 입력 방법]", 11, True)
    r = add_line(r, "ID      : B열에 'ID' 선택 -> D열(num)은 자동으로 1씩 증가 -> F열에 cycle 숫자를 직접 입력")
    r = add_line(r, "Power   : B열에 'Power' 선택 -> C열에 'command' 자동표시 -> D열 드롭다운에서 명령어 선택")
    r = add_line(r, "          (ACC_IGN_ON, ACC_IGN_OFF, ACC_ON, IGN_ON, ACC_OFF, IGN_OFF)")
    r = add_line(r, "delay   : B열에 'delay' 선택 -> C열에 'ms' 자동표시 -> D열에 숫자(ms)를 직접 입력")
    r = add_line(r, "CANReq  : B열에 'CANReq' 선택 후 I열(하늘색)에서 Signal 검색·선택 -> F열(Signal)·D열(Message) 자동 채움")
    r = add_line(r, "          또는 F열 드롭다운에서 직접 선택/타이핑 -> D열(Message) 자동 채움 (검색 없음)")
    r = add_line(r, "          H열에 Value를 직접 입력하세요 (K열에 DBC 참고값이 자동 표시됨)")
    r = add_line(r, "Audio   : B열에 'Audio' 선택 -> C열에 'command' 자동표시 -> D열 드롭다운에서 명령어 선택")
    r = add_line(r, "          (StartREC, StopREC, compWAV, RMSmeasure)")
    r = add_line(r, "CANEv   : CANReq와 동일한 방식으로 입력")
    r = add_line(r, "loop    : B열에 'loop' 선택 -> D열에 반복 횟수(cycle)를 직접 입력 -> F열에 '[' 자동 표시")
    r += 1
    r = add_line(r, "[CAN DB 참고값 (J열)]", 11, True)
    r = add_line(r, "CANReq/CANEv 행에서 F열에 신호명을 입력하면, 해당 신호에 DBC(VAL_)로 정의된 값들이")
    r = add_line(r, "'16진수값=설명' 형태로 J열에 자동 표시됩니다. H열에 값을 입력할 때 참고하세요.")
    r = add_line(r, f"현재 DB에는 신호 {n_sig}개, VAL_ 정의값 {n_val}개가 들어있습니다 (원본: {dbc_filename}).")
    r += 1
    r = add_line(r, "[DBC가 개정되었을 때]", 11, True)
    r = add_line(r, "이 스크립트를 새 .dbc 파일로 다시 실행하면, 동일한 형식의 파일이 새로 생성됩니다")
    r = add_line(r, "(DB / DB_VAL 시트 데이터만 새로 반영되고 수식/서식은 동일하게 유지됩니다).")
    r += 1
    r = add_line(r, "[주의사항]", 11, True)
    r = add_line(r, "- 드롭다운 목록에 없는 값을 입력하면 '계속하시겠습니까?' 경고가 뜰 수 있습니다.")
    r = add_line(r, "  숫자를 직접 입력하는 칸(delay/loop/ID의 cycle)에서는 무시하고 '예'를 누르면 됩니다.")
    r = add_line(r, "- 행의 B열(Step 종류)을 변경하면 이전에 입력했던 D/F/H열 값은 자동으로 지워지지 않으므로")
    r = add_line(r, "  직접 확인 후 정리해 주세요.")
    r = add_line(r, "- 동일한 이름의 CAN 신호가 여러 메시지에 존재하면, D열(Message)은 첫 번째로 검색된")
    r = add_line(r, "  메시지만 표시합니다.")
    r = add_line(r, "- CANReq/CANEv: I열에서 Signal을 선택하면 F열(Signal)과 D열(Message)이 자동으로 채워집니다.")
    r = add_line(r, "  F열은 수식 셀이므로 직접 편집하면 I열과 연동이 깨질 수 있습니다 (I열을 지우면 F열도 초기화됩니다).")
    r = add_line(r, "  검색 없이 직접 입력할 경우 F열 드롭다운을 사용하면 되고, 이때 I열은 비워두세요.")

    wb.move_sheet("사용법", offset=-4)

    wb.save(out_path)

    # openpyxl 버그 우회: dxf(조건부 서식) fill에 patternType이 누락되어
    # Excel이 "파일에 문제가 있습니다" 오류를 표시함 → styles.xml을 직접 패치
    import zipfile as _zf, os as _os, re as _re
    _tmp = out_path + ".tmp.xlsx"
    with _zf.ZipFile(out_path, "r") as _zin, \
         _zf.ZipFile(_tmp, "w", _zf.ZIP_DEFLATED) as _zout:
        for _item in _zin.infolist():
            _data = _zin.read(_item.filename)
            if _item.filename == "xl/styles.xml":
                _xml = _data.decode("utf-8")
                # dxfs 섹션 내 patternType 없는 <patternFill>에만 추가
                _xml = _re.sub(
                    r'(<dxfs[^>]*>.*?)(<patternFill>)(<[bf]gColor)',
                    lambda m: m.group(1) + '<patternFill patternType="solid">' + m.group(3),
                    _xml, flags=_re.DOTALL
                )
                _data = _xml.encode("utf-8")
            _zout.writestr(_item, _data)
    _os.replace(_tmp, out_path)

    return n_sig, n_val


# ---------------------------------------------------------------------------
# 3. CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="CAN DBC 파일을 읽어서 테스트 스크립트 입력용 엑셀 파일을 생성합니다."
    )
    parser.add_argument("dbc_path", help="입력 DBC 파일 경로 (.dbc)")
    parser.add_argument("xlsx_path", nargs="?", default=None,
                         help="출력 엑셀 파일 경로 (기본값: DBC 파일명 + _ScriptEditor.xlsx)")
    args = parser.parse_args()

    out_path = args.xlsx_path
    if out_path is None:
        out_path = re.sub(r"\.dbc$", "", args.dbc_path, flags=re.IGNORECASE) + "_ScriptEditor.xlsx"

    print(f"[1/3] DBC 파싱 중: {args.dbc_path}")
    sig_rows, messages, duplicates, val_entries = parse_dbc(args.dbc_path)
    print(f"      메시지 {len(messages)}개, 신호 {len(sig_rows)}개, VAL_ 정의값 {len(val_entries)}개 추출 완료")

    if duplicates:
        print(f"[!] 주의: 이름이 중복된 신호 {len(duplicates)}개 발견 (조회 시 첫 번째 메시지만 표시됨):")
        for sig, msgs in list(duplicates.items())[:10]:
            print(f"      - {sig}: {msgs}")
        if len(duplicates) > 10:
            print(f"      ... 외 {len(duplicates) - 10}개")

    print(f"[2/3] 엑셀 파일 생성 중: {out_path}")
    dbc_filename = args.dbc_path.split("/")[-1].split("\\")[-1]
    n_sig, n_val = build_workbook(sig_rows, val_entries, dbc_filename, out_path)

    print(f"[3/3] 완료: {out_path}  (신호 {n_sig}개 / VAL_ 정의값 {n_val}개)")


if __name__ == "__main__":
    main()
